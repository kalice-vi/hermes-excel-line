"""brain_store.py — excel_line v2: cây phân cấp, mỗi file .xlsx tối đa 10 dòng.

QUY ƯỚC (bất biến, code cưỡng chế):
  brain.xlsx          : gốc (layer 1) — header + TỐI ĐA 10 dòng
  6 cột               : ID (int) | Title (≤50) | Content (≤250) | Tags
                        | Branch (đường dẫn tương đối root) | Updated
  Branch .xlsx        : ĐIỂM NHÁNH — trỏ tới file hàng khác (nội bộ)
  Branch non-.xlsx    : LÁ — file tài nguyên (.py/.png/.md/…), bắt buộc nằm
                        cuối cây; thư mục chứa lá = tên file gốc ('skill.xlsx'
                        → tài nguyên trong folder 'skill/')
  Branch rỗng         : memory thuần không kèm file
  Sau một đối tượng mà còn ≥1 đối tượng khác → đối tượng đó phải là .xlsx
  (promote). Không có ngoại lệ.

  Đầy 10 dòng → add raise FullError: LLM phải merge (nén vào memory cũ)
  hoặc child+move (tách nhánh) — không im lặng ghi tràn.
"""
from __future__ import annotations
import os, re, time, json
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

try:
    from .store import ExcelLineStore, _now
except ImportError:
    from store import ExcelLineStore, _now  # standalone / single-file test run

_safe_cell = ExcelLineStore._safe_cell

MASTER_V2 = "brain.xlsx"
COLS = ["ID", "Title", "Content", "Tags", "Branch", "Updated"]
MAX_ROWS = 10
TITLE_CAP, CONTENT_CAP = 50, 250


class FullError(Exception):
    def __init__(self, branch):
        super().__init__(f"{branch} đầy {MAX_ROWS}/{MAX_ROWS} — merge hoặc child+move")
        self.branch = branch


class BadBranch(Exception):
    pass


class BrainStore(ExcelLineStore):
    """Kế thừa lock đa tiến trình của v1, đổi toàn bộ layout sang cây v2.

    KHÔNG dùng master phẳng v1: _ensure_master/_next_seq bị ghi đè và file
    excel-line_index.xlsx không còn vai trò gì (trên Windows handle mở giữ
    khóa xóa file → phải né mọi code path v1 chạm tới nó)."""

    def __init__(self, root_dir: str):
        self.root = root_dir
        os.makedirs(self.root, exist_ok=True)
        self._lock = __import__("threading").RLock()
        self._seq = 0
        self._master = os.path.join(root_dir, MASTER_V2)   # v2: brain.xlsx
        self._lockfile = os.path.join(root_dir, ".excel_line.lock")
        self._ensure(self._master)

    # ---------------- io ----------------
    def _ensure(self, path: str):
        if not os.path.exists(path):
            wb = Workbook(); ws = wb.active; ws.title = "mem"
            ws.append(COLS)
            for i in range(1, 7):
                ws.column_dimensions[get_column_letter(i)].width = 28
            wb.save(path)

    def resolve(self, branch: str) -> str:
        """branch → đường dẫn tuyệt đối.

        HAI loại:
          - nội bộ (rel-root, ví dụ 'user.xlsx', 'skill/x.py'): nằm trong
            thư mục excel_line — asset do plugin tạo thì để ở đây.
          - LÁ BÊN NGOÀI (tuyệt đối, ví dụ
            'C:/Users/Admin/AppData/Local/hermes/skills/foo/SKILL.md' hoặc
            'D:/YTECH.../Format.xlsx'): file KHÔNG di chuyển — cây chỉ giữ
            "cuống lá" trỏ tới. Khuyến nghị của user 03/09: lá tự do bay
            lượn nơi nó vốn sống; branch là đường dẫn truy xuất.
        """
        b = (branch or "").strip().replace("\\", "/").lstrip("/")
        if not b or ".." in b.split("/") and ":" not in b:
            raise BadBranch("branch không hợp lệ: " + repr(branch))
        # absolute windows/posix path => external leaf (read/write pointer only)
        if re.match(r"^[A-Za-z]:/", b) or b.startswith("/"):
            return os.path.normpath(b)
        fp = os.path.normpath(os.path.join(self.root, b))
        if not fp.startswith(os.path.normpath(self.root)):
            raise BadBranch("branch vượt ra ngoài root")
        return fp

    def is_external(self, branch: str) -> bool:
        return bool(re.match(r"^([A-Za-z]:/|/)", (branch or "").replace("\\", "/")))

    def load_rows(self, branch: str) -> List[Dict]:
        fp = self.resolve(branch)
        if not os.path.exists(fp):
            raise BadBranch("không có file: " + branch)
        try:
            wb = load_workbook(fp, read_only=True, data_only=True)
            try:
                ws = wb[wb.sheetnames[0]]
                header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
                rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)
                        if r and r[0] is not None]
            finally:
                wb.close()
        except Exception:
            # read_only giữ file handle trên Windows → writer khác có thể đang
            # giữ khóa; thử lại ngay bằng chế độ thông thường (copy vào memory)
            import zipfile
            with open(fp, "rb") as f:
                data = f.read()
            import io
            wb = load_workbook(io.BytesIO(data))
            ws = wb[wb.sheetnames[0]]
            header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
            rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)
                    if r and r[0] is not None]
            wb.close()
        return rows

    def _write_rows(self, fp: str, rows: List[Dict]):
        tmp = fp + ".tmp"
        wb = Workbook(); ws = wb.active; ws.title = "mem"
        ws.append(COLS)
        for r in rows:
            ws.append([r.get("id"), _safe_cell(str(r.get("title") or ""))[:TITLE_CAP + 5],
                       _safe_cell(str(r.get("content") or ""))[:CONTENT_CAP + 10],
                       _safe_cell(str(r.get("tags") or "")),
                       _safe_cell(str(r.get("branch") or "")),
                       r.get("updated") or _now()])
        for i in range(1, 7):
            ws.column_dimensions[get_column_letter(i)].width = 28
        wb.save(tmp)
        os.replace(tmp, fp)

    def _all_ids(self) -> List[int]:
        ids = []
        for dp, _dn, fns in os.walk(self.root):
            for f in fns:
                if f.endswith(".xlsx") and not f.endswith(".bak") and "_v1" not in dp:
                    try:
                        ids += [int(r["id"]) for r in self.load_rows(
                            os.path.relpath(os.path.join(dp, f), self.root).replace("\\", "/"))
                            if isinstance(r.get("id"), int)]
                    except Exception:
                        pass
        return ids

    def next_id(self) -> int:
        """ID không BAO GIỜ lặp lại, kể cả sau khi xóa dòng: cao-trịch-lưu
        (.id_seq high-water) là chuẩn, kết hợp max thực tế để an toàn.
        (Chính bug tái-use id khi writer mới đọc max cũ là nguồn 80 cặp trùng
        id của v1 — v2 cắt tận gốc.)"""
        seq_fp = os.path.join(self.root, ".id_seq")
        last = 0
        try:
            with open(seq_fp, encoding="utf-8") as f:
                last = int(json.load(f).get("last", 0))
        except Exception:
            last = 0
        rid = max([last] + self._all_ids()) + 1
        try:
            tmp = seq_fp + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump({"last": rid}, f)
            os.replace(tmp, seq_fp)
        except Exception:
            pass
        return rid

    # ---------------- validation ----------------
    def _validate_branch_value(self, link: str):
        if not link:
            return
        norm = link.replace("\\", "/")
        if norm.endswith("/"):
            raise BadBranch("branch phải là FILE, không phải thư mục: " + link)
        if self.is_external(norm):
            # LÁ BÊN NGOÀI: trỏ tới file nơi nó vốn sống (skills Hermes,
            # thư mục user…). Không di chuyển, không tạo node .xlsx ngoài root.
            if norm.lower().endswith(".xlsx"):
                raise BadBranch("node .xlsx phải nằm TRONG cây excel_line: " + link)
            if not os.path.exists(norm):
                raise BadBranch("lá ngoài trỏ tới file chưa tồn tại: " + link)
            return
        if norm.lower().endswith(".xlsx"):
            fp = self.resolve(link)
            if not os.path.exists(fp):
                raise BadBranch("node trỏ tới file chưa tồn tại: " + link
                                + " (dùng child() để tạo trước)")
        else:
            fp = self.resolve(link)
            if not os.path.exists(fp):
                raise BadBranch("lá trỏ tới file chưa tồn tại: " + link)

    # ---------------- API ----------------
    def add(self, branch: str = MASTER_V2, title: str = "", content: str = "",
            tags: str = "", link: str = "", **compat) -> int:
        """Thêm dòng vào file node. Đầy → FullError (ép LLM quyết định).

        Compat: gọi kiểu cũ add(zone=..., brief=...) → route xuống nhánh
        <zone>.xlsx tương ứng (tự tạo nếu chưa có)."""
        if "zone" in compat or "brief" in compat:
            zone = str(compat.get("zone") or "knowledge")
            br = str(compat.get("brief") or "")[:250]
            ti = str(compat.get("title") or br)[:50]
            co = str(compat.get("content") or br)[:250]
            tg = str(compat.get("tags") or "")
            return self._compat_zone_add(zone, br, co, ti, tg)
        branch = branch or MASTER_V2
        with self._cross_process_lock(timeout=15.0):
            rows = self.load_rows(branch)
            if len(rows) >= MAX_ROWS:
                raise FullError(branch)
            self._validate_branch_value(link)
            rid = self.next_id()
            rows.append({"id": rid, "title": (title or "")[:TITLE_CAP],
                         "content": (content or "")[:CONTENT_CAP],
                         "tags": tags or "", "branch": link or "",
                         "updated": _now()})
            self._write_rows(self.resolve(branch), rows)
            return rid

    # --- v1 shim: worker dùng store.add(zone=..., brief=...) ---
    def _compat_zone_add(self, zone, brief, content, title, tags):
        try:
            b = f"{zone}.xlsx"
            if not os.path.exists(self.resolve(b)):
                return self.child(MASTER_V2, zone, (title or brief)[:TITLE_CAP],
                                  content[:CONTENT_CAP], tags)["node_id"]
            return self.add(b, (title or brief)[:TITLE_CAP], content[:CONTENT_CAP], tags)
        except FullError:
            return -1
        except Exception:
            return -1

    def v1_add(self, zone: str, brief: str, content: str, title: str = "",
               tags: str = "") -> int:
        return self._compat_zone_add(zone or "knowledge", brief or "", content or "",
                                     title or "", tags or "")

    def child(self, parent: str, name: str, title: str, content: str = "",
              tags: str = "") -> Dict:
        """Tạo điểm nhánh mới: <name>.xlsx + một dòng trỏ tới nó trong parent."""
        name = name.strip().replace("\\", "/")
        if not name or "/" in name or not name.isascii():
            # cho phép ascii alphanumeric + -_
            pass
        node_branch = (name if name.endswith(".xlsx") else name + ".xlsx")
        # nếu name nằm trong folder con: giữ nguyên; name bare → folder theo stem parent
        parent_fp = self.resolve(parent)
        stem = os.path.splitext(os.path.basename(parent_fp))[0]
        node_rel = node_branch if parent == MASTER_V2 else f"{stem}/{node_branch}"
        node_fp = self.resolve(node_rel)
        with self._cross_process_lock(timeout=15.0):
            os.makedirs(os.path.dirname(node_fp), exist_ok=True)
            self._ensure(node_fp)
            rows = self.load_rows(parent)
            if len(rows) >= MAX_ROWS:
                raise FullError(parent)
            nid = self.next_id()
            rows.append({"id": nid, "title": (title or name)[:TITLE_CAP],
                         "content": (content or "")[:CONTENT_CAP],
                         "tags": tags or "", "branch": node_rel, "updated": _now()})
            self._write_rows(parent_fp, rows)
            return {"node_id": nid, "file": node_rel}

    def move(self, src: str, ids: List[int], dst: str) -> int:
        """Chuyển dòng giữa các file node (dst phải còn chỗ)."""
        with self._cross_process_lock(timeout=15.0):
            srows = self.load_rows(src)
            drows = self.load_rows(dst)
            picked = [r for r in srows if r["id"] in set(ids)]
            if not picked:
                raise BadBranch("không tìm thấy id trong " + src)
            if len(drows) + len(picked) > MAX_ROWS:
                raise FullError(dst)
            for r in picked:
                r["updated"] = _now()
                srows.remove(r)
                drows.append(r)
            self._write_rows(self.resolve(src), srows)
            self._write_rows(self.resolve(dst), drows)
            return len(picked)

    def merge(self, branch: str, ids: List[int], title: str, content: str,
              tags: str = "") -> int:
        """Nén nhiều dòng cũ thành MỘT memory mới (giải phóng dung lượng)."""
        with self._cross_process_lock(timeout=15.0):
            rows = self.load_rows(branch)
            picked = [r for r in rows if r["id"] in set(ids)]
            if len(picked) < 1:
                raise BadBranch("không có dòng nào để merge")
            keep_branch = picked[0].get("branch") or ""
            for r in picked:
                rows.remove(r)
            rid = self.next_id()
            rows.append({"id": rid, "title": (title or "")[:TITLE_CAP],
                         "content": (content or "")[:CONTENT_CAP],
                         "tags": tags or "", "branch": keep_branch,
                         "updated": _now()})
            self._write_rows(self.resolve(branch), rows)
            return rid

    def set(self, branch: str, row_id: int, title: str = "", content: str = "",
            tags: str = "", link: Optional[str] = None) -> bool:
        """Sửa字段 của một dòng (map rename / retag dùng path này)."""
        with self._cross_process_lock(timeout=15.0):
            rows = self.load_rows(branch)
            for r in rows:
                if r["id"] == row_id:
                    if title:
                        r["title"] = title[:TITLE_CAP]
                    if content:
                        r["content"] = content[:CONTENT_CAP]
                    if tags is not None and tags != "":
                        r["tags"] = tags
                    if link is not None:
                        self._validate_branch_value(link)
                        r["branch"] = link
                    r["updated"] = _now()
                    self._write_rows(self.resolve(branch), rows)
                    return True
            return False

    def rm(self, branch: str, row_id: int, purge_file: bool = False) -> bool:
        with self._cross_process_lock(timeout=15.0):
            rows = self.load_rows(branch)
            for r in rows:
                if r["id"] == row_id:
                    rows.remove(r)
                    self._write_rows(self.resolve(branch), rows)
                    if purge_file and r.get("branch") and not r["branch"].lower().endswith(".xlsx"):
                        if self.is_external(r["branch"]):
                            pass  # LÁ BÊN NGOÀI: không bao giờ xóa/dời file nơi nó sống
                        else:
                            try:
                                os.remove(self.resolve(r["branch"]))
                            except OSError:
                                pass
                    return True
            return False

    def promote(self, branch: str, row_id: int, name: str) -> Dict:
        """LÁ .py đang lớn thành phần mềm → thay chỗ bằng .xlsx đẩy file ra sau.

        Result: cùng vị trí dòng giờ là node <name>.xlsx chứa 1 dòng lá trỏ tới
        asset cũ (asset được dời vào folder '<name>/').
        Thứ tự: XÓA dòng lá cũ TRƯỚC (giữ slot cha) → tạo node → dời asset →
        ghi lá vào node. Net slot của cha không đổi nên promote luôn hợp lệ."""
        rows = self.load_rows(branch)
        orig = next((r for r in rows if r["id"] == row_id), None)
        if orig is None:
            raise BadBranch("không thấy dòng #" + str(row_id) + " trong " + branch)
        asset = str(orig.get("branch") or "")
        self.rm(branch, row_id)                    # giải phóng slot
        node = self.child(branch, name, title=str(orig.get("title") or name)[:TITLE_CAP],
                          content=str(orig.get("content") or ""),
                          tags=str(orig.get("tags") or ""))
        node_fp = self.resolve(node["file"])
        new_asset_rel = ""
        if asset:
            if self.is_external(asset):
                # LÁ BÊN NGOÀI: giữ nguyên chỗ ở, chỉ mang theo cuống lá
                new_asset_rel = asset
            else:
                src_abs = self.resolve(asset)
                dst_rel = os.path.join(os.path.basename(node["file"]).replace(".xlsx", ""),
                                       os.path.basename(asset)).replace("\\", "/")
                dst_abs = self.resolve(dst_rel)
                os.makedirs(os.path.dirname(dst_abs), exist_ok=True)
                if src_abs != dst_abs:
                    shutil_move(src_abs, dst_abs)
                new_asset_rel = os.path.relpath(dst_abs, self.root).replace("\\", "/")
            # dòng lá trong node mới
            self.add(node["file"], title=str(orig.get("title") or "")[:TITLE_CAP],
                     content=str(orig.get("content") or "")[:CONTENT_CAP],
                     tags=str(orig.get("tags") or ""), link=new_asset_rel)
        return node

    # ---------------- đọc cây ----------------
    def tree(self) -> Dict:
        """Cây đệ quy từ gốc — dùng cho map và prompt."""
        def walk(branch: str) -> Dict:
            node = {"branch": branch, "rows": [], "children": {}}
            try:
                rows = self.load_rows(branch)
            except Exception:
                return node
            for r in rows:
                rb = str(r.get("branch") or "")
                if rb.lower().endswith(".xlsx") and rb:
                    node["children"][rb] = walk(rb)
                    r["_kind"] = "node"
                elif rb:
                    r["_kind"] = "leaf"
                else:
                    r["_kind"] = "mem"
                node["rows"].append(r)
            return node
        return walk(MASTER_V2)

    def tree_text(self) -> str:
        out = []
        def render(n: Dict, depth: int):
            usage = f" ({len(n['rows'])}/{MAX_ROWS})" if depth else ""
            out.append("  " * depth + n["branch"] + usage)
            for r in n["rows"]:
                if r["_kind"] != "node":
                    mark = "🔗" if r["_kind"] == "leaf" else "  "
                    out.append("  " * (depth + 1) + f"#{r['id']} {mark}{str(r['title'])[:44]}"
                               + (f" → {r['branch']}" if r["branch"] and r["_kind"] == "leaf" else ""))
            for ch in n["children"].values():
                render(ch, depth + 1)
        render(self.tree(), 0)
        return "\n".join(out)

    def find(self, row_id: int) -> Optional[Dict]:
        """Tìm dòng theo id toàn cây → {branch, row}."""
        def walk(b):
            for r in self.load_rows(b):
                if r["id"] == row_id:
                    return {"branch": b, "row": r}
                rb = str(r.get("branch") or "")
                if rb.lower().endswith(".xlsx"):
                    got = walk(rb)
                    if got:
                        return got
            return None
        return walk(MASTER_V2)

    def search(self, q: str, limit: int = 20) -> List[Dict]:
        q = (q or "").lower().strip()
        hits = []
        def walk(b):
            for r in self.load_rows(b):
                blob = " ".join(str(r.get(k) or "") for k in ("title", "content", "tags")).lower()
                if q and q in blob:
                    hits.append({"branch": b, "path": self.path_of(r["id"]),
                                 "id": r["id"], "title": r["title"],
                                 "tags": r["tags"]})
                rb = str(r.get("branch") or "")
                if rb.lower().endswith(".xlsx"):
                    walk(rb)
        if q:
            try:
                walk(MASTER_V2)
            except BadBranch:
                pass
        return hits[:limit]

    def path_of(self, row_id: int) -> str:
        """Đường dẫn cây tới một id: brain.xlsx › user.xlsx › #17
        (danh sách FILE node, kết thúc bằng id của chính dòng)."""
        chain = []
        def walk(b, acc):
            for r in self.load_rows(b):
                rb = str(r.get("branch") or "")
                nxt = acc + [b]
                if r["id"] == row_id:
                    chain.extend(nxt)
                    return True
                if rb.lower().endswith(".xlsx"):
                    if walk(rb, nxt):
                        return True
            return False
        walk(MASTER_V2, [])
        if not chain:
            f = self.find(row_id)
            return f["branch"] if f else ""
        return " › ".join(chain) + f" › #{row_id}"


import shutil as _shutil
def shutil_move(a, b):
    _shutil.move(a, b)


# ------------------------------------------------------------------ v1 surface
# BrainStore additionally answers the OLD ExcelLineStore API so the existing
# provider/worker code keeps working on the new tree (read_zone/update/delete/
# forget/search_index/count/list_zones/zone_path).

def _flatten_rows(store) -> List[Dict]:
    out = []
    def walk(b):
        for r in store.load_rows(b):
            out.append({**r, "_branch": b})
            rb = str(r.get("branch") or "")
            if rb.lower().endswith(".xlsx"):
                walk(rb)
    walk(MASTER_V2)
    return out

def search_index(self, query: str, limit: int = 10) -> List[Dict]:
    q = (query or "").lower().strip()
    if not q:
        return []
    hits = []
    for r in _flatten_rows(self):
        blob = " ".join(str(r.get(k) or "") for k in ("title", "content", "tags", "branch")).lower()
        if q in blob:
            hits.append({"id": r["id"], "zone": os.path.splitext(r["_branch"])[0],
                         "brief": r.get("title") or str(r.get("content") or "")[:120],
                         "title": r.get("title"), "path": self.zone_path(
                             os.path.splitext(r["_branch"])[0]),
                         "tags": r.get("tags") or ""})
    return hits[:limit]

def read_zone(self, path: str, limit: int = 20) -> List[Dict]:
    try:
        rel = os.path.relpath(path, self.root).replace("\\", "/")
        rows = self.load_rows(rel)
    except Exception:
        return []
    return [{"id": r["id"], "title": r.get("title"), "content": r.get("content"),
             "tags": r.get("tags"), "created": r.get("updated"),
             "branch": r.get("branch")} for r in rows[-limit:]]

def zone_path(self, zone: str) -> str:
    z = str(zone or "").strip()
    if z.endswith(".xlsx"):
        return os.path.join(self.root, z)
    return os.path.join(self.root, f"{z}.xlsx")

def list_zones(self) -> List[str]:
    out = []
    for dp, _dn, fns in os.walk(self.root):
        for f in fns:
            if f.endswith(".xlsx") and not f.endswith(".tmp"):
                out.append(f[:-5])
    return sorted(set(out))

def count(self) -> int:
    try:
        return len(_flatten_rows(self))
    except Exception:
        return 0

def update(self, zone: str = "", row_id: int = 0, brief: str = "", content: str = "",
           title: str = "", tags: str = "") -> bool:
    f = self.find(int(row_id)) if row_id else None
    if not f:
        return False
    return self.set(f["branch"], int(row_id),
                    title=(title or brief)[:TITLE_CAP], content=content, tags=tags)

def delete(self, zone: str = "", row_id: int = 0) -> bool:
    f = self.find(int(row_id)) if row_id else None
    if not f:
        return False
    return self.rm(f["branch"], int(row_id))

def forget(self, query: str) -> int:
    q = (query or "").lower().strip()
    if not q:
        return 0
    rows = [r for r in _flatten_rows(self)
            if q in " ".join(str(r.get(k) or "") for k in ("title", "content", "tags")).lower()]
    n = 0
    for r in rows:
        if self.rm(r["_branch"], r["id"]):
            n += 1
    return n

BrainStore.search_index = search_index
BrainStore.read_zone = read_zone
BrainStore.zone_path = zone_path
BrainStore.list_zones = list_zones
BrainStore.count = count
BrainStore.update = update
BrainStore.delete = delete
BrainStore.forget = forget
