"""store.py — persistence layer for the excel_line memory provider.

INTRODUCTION
    This module owns all Excel read/write for excel_line. It is the single
    source of truth for durable memory: it appends brief rows to a MASTER index
    and detailed records to per-zone workbooks, and answers search/read queries.
    It contains no agent-runtime imports, so it can run standalone (unit tests,
    the sub-agent worker) without loading the full provider.

LAYOUT
    - MASTER workbook (excel-line_index.xlsx)
        Sheet "index": one row per memory entry
            A id | B zone | C brief | D title | E path | F tags | G created | H updated
    - ZONE workbooks (<zone>.xlsx) e.g. user.xlsx, project.xlsx, knowledge.xlsx
        Sheet "mem": A id | B title | C content | D tags | E created | F updated

    All workbooks live under one root dir (default $HERMES_HOME/excel_line/).
    The master index keeps only a brief note + title + the path to the detailed
    zone record, so the index stays small and human-scannable while deep
    knowledge lives in per-zone files.

THREAD / PROCESS SAFETY
    Writes are guarded by an in-process RLock AND a cross-process file lock
    (atomic O_EXCL with pid-checked stale-lock recovery), so concurrent writes
    from the provider and the background worker never corrupt the workbooks.
"""

from __future__ import annotations

import os
import threading
import time
import contextlib
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

MASTER_NAME = "excel-line_index.xlsx"
INDEX_COLS = ["id", "zone", "brief", "title", "path", "tags", "created", "updated"]
MEM_COLS = ["id", "title", "content", "tags", "created", "updated"]

ZONE_DEFAULTS = ["user", "project", "pref", "task", "knowledge", "contact", "skill"]


def _now() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


class ExcelLineStore:
    """Manages the master index and per-zone workbooks.

    Uses BOTH an in-process RLock (for thread safety) and a cross-process
    file lock (for safety when multiple Hermes processes share one `root`),
    so concurrent writes from two processes cannot corrupt the workbooks.
    """

    def __init__(self, root_dir: str):
        self.root = root_dir
        os.makedirs(self.root, exist_ok=True)
        self._lock = threading.RLock()
        self._seq = 0
        self._master = os.path.join(self.root, MASTER_NAME)
        self._lockfile = os.path.join(self.root, ".excel_line.lock")
        self._ensure_master()
        self._seq = self._next_seq()

    # -- cross-process lock ------------------------------------------------
    @contextlib.contextmanager
    def _cross_process_lock(self, timeout: float = 10.0, stale_after: float = 30.0):
        """Atomic file-lock usable across processes (works on Win + POSIX).

        The lock file stores ``<pid>:<timestamp>``. A stale lock (older than
        `stale_after`, or owned by a process that no longer exists) is cleared
        so we never block forever and never collide with a live owner.
        """
        deadline = time.time() + timeout
        fd = None
        while fd is None:
            try:
                fd = os.open(self._lockfile, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                # Record ownership so a later process can tell if we're alive.
                os.write(fd, f"{os.getpid()}:{time.time()}".encode())
                break
            except OSError:
                # Maybe a stale lock from a crashed process — try to clear it.
                try:
                    age = time.time() - os.path.getmtime(self._lockfile)
                    if age > stale_after:
                        # Also confirm the recorded pid is dead (if readable).
                        try:
                            with open(self._lockfile, "r") as lf:
                                owner = lf.read().split(":", 1)[0].strip()
                            if owner and owner.isdigit():
                                import signal
                                try:
                                    os.kill(int(owner), 0)  # alive?
                                    break  # live owner -> give up this round
                                except OSError:
                                    pass  # dead -> safe to remove
                        except Exception:
                            pass
                        os.remove(self._lockfile)
                        continue
                except OSError:
                    pass
                if time.time() > deadline:
                    # FAIL-CLOSED: do NOT proceed without the lock. Yielding with
                    # fd=None would let the caller run its critical section
                    # unsynchronized, causing duplicate IDs / corrupted workbooks
                    # under cross-process contention (ChatGPT review B1).
                    raise TimeoutError(
                        "excel_line: could not acquire cross-process lock within "
                        f"{timeout}s; aborting to avoid unsynchronized write")
                time.sleep(0.05)
        if fd is None:
            raise TimeoutError("excel_line: cross-process lock not held")
        try:
            yield
        finally:
            if fd is not None:
                try:
                    os.close(fd)
                except OSError:
                    pass
                # Only remove the lock file if it is STILL ours. A concurrent
                # process may have taken over the lock (and rewritten the file
                # with its own pid) while we were inside `yield`, so blindly
                # os.remove() here would delete the *other* process's lock and
                # invalidate their mutual exclusion (Gemini review BLOCKER-02).
                try:
                    with open(self._lockfile, "r") as lf:
                        owner = lf.read().split(":", 1)[0].strip()
                    if owner == str(os.getpid()):
                        os.remove(self._lockfile)
                except (OSError, ValueError):
                    # File gone or unreadable — another process already cleared
                    # it; nothing to do.
                    pass

    # -- paths -------------------------------------------------------------

    def zone_path(self, zone: str) -> str:
        safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in zone.lower())
        return os.path.join(self.root, f"{safe}.xlsx")

    # -- master ------------------------------------------------------------

    # -- formula-injection safety -----------------------------------------
    @staticmethod
    def _safe_cell(value: str) -> str:
        """Prevent Excel formula injection: a cell whose text starts with
        = + - @ is interpreted as a formula when the workbook is opened in
        Excel (and can trigger DDE/hyperlink command execution). Prefix such
        values with an apostrophe so Excel treats them as literal text."""
        s = "" if value is None else str(value)
        if s and s[0] in "=+-@\t\r":
            return "'" + s
        return s

    def _ensure_master(self):
        if not os.path.exists(self._master):
            wb = Workbook()
            ws = wb.active
            ws.title = "index"
            ws.append(INDEX_COLS)
            self._autosize(ws, INDEX_COLS)
            wb.save(self._master)

    def _next_seq(self) -> int:
        with self._cross_process_lock():
            try:
                wb = load_workbook(self._master, read_only=True)
                ws = wb["index"]
                max_id = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and isinstance(row[0], int):
                        max_id = max(max_id, row[0])
                return max_id
            except Exception:
                return 0

    def _new_id(self) -> int:
        # Compute the next id from the persisted master index (not an in-memory
        # counter) so that concurrent ExcelLineStore instances in different
        # processes never hand out the same id (ChatGPT review B2). The caller
        # must already hold self._lock (and usually the cross-process lock too).
        max_id = 0
        try:
            wb = load_workbook(self._master, read_only=True)
            ws = wb["index"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and isinstance(row[0], int):
                    max_id = max(max_id, row[0])
        except Exception:
            max_id = 0
        return max_id + 1

    # -- write -------------------------------------------------------------

    def add(self, zone: str, brief: str, content: str, title: str = "",
            tags: str = "") -> int:
        """Add a memory: a brief row in the master index + a detailed zone row.

        Wrapped in try/except so a transient file/lock error cannot crash the
        provider — the caller gets -1 and a logged error instead of an exception.
        """
        zone = zone or "knowledge"
        # W1: enforce zone whitelist at the public store boundary, not only in
        # the worker classifier — otherwise the agent could create arbitrary
        # .xlsx zone workbooks via the direct add() API.
        if zone not in ZONE_DEFAULTS:
            try:
                logger.debug("excel_line add rejected unknown zone: %s", zone)
            except Exception:
                pass
            return -1
        try:
            with self._cross_process_lock():
                with self._lock:
                    mid = self._new_id()
                    now = _now()
                    zpath = self.zone_path(zone)
                    # write zone detail
                    if not os.path.exists(zpath):
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "mem"
                        ws.append(MEM_COLS)
                        self._autosize(ws, MEM_COLS)
                    else:
                        wb = load_workbook(zpath)
                        ws = wb["mem"]
                    zid = mid  # CRITICAL: zone row id MUST equal the master id,
                               # otherwise update/delete/forget (which receive the
                               # master id from add()) cannot locate the zone row.
                    ws.append([zid, self._safe_cell(title or brief[:40]),
                               self._safe_cell(content), self._safe_cell(tags), now, now])
                    wb.save(zpath)
                    # write master index row (only after zone committed successfully)
                    try:
                        mwb = load_workbook(self._master)
                        mws = mwb["index"]
                        mws.append([mid, zone, self._safe_cell(brief),
                                    self._safe_cell(title), zpath, self._safe_cell(tags), now, now])
                        mwb.save(self._master)
                    except Exception:
                        # Roll back the zone row we just appended so zone and
                        # master never diverge (ChatGPT review B3: non-atomic add).
                        # The caller still gets -1 and a logged error.
                        try:
                            wb2 = load_workbook(zpath)
                            ws2 = wb2["mem"]
                            if ws2.max_row >= 2:
                                ws2.delete_rows(ws2.max_row, 1)
                            wb2.save(zpath)
                        except Exception:
                            pass
                        raise
                    return mid
        except Exception as e:  # pragma: no cover - defensive
            import logging
            logging.getLogger(__name__).error("excel_line store.add failed: %s", e)
            return -1

    # -- read --------------------------------------------------------------

    def search_index(self, query: str, limit: int = 10) -> List[Dict]:
        """Keyword scan over the master index (brief + tags + zone)."""
        q = query.lower()
        hits: List[Dict] = []
        with self._cross_process_lock():  # B4: reads must also exclude writers
            with self._lock:
                wb = load_workbook(self._master, read_only=True)
                ws = wb["index"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    # Scan zone + brief + title (row[3]) + path + tags.
                    blob = " ".join(str(x or "") for x in (row[1], row[2], row[3], row[4], row[5])).lower()
                    if q in blob:
                        hits.append({
                            "id": row[0], "zone": row[1], "brief": row[2],
                            "title": row[3], "path": row[4], "tags": row[5],
                        })
        return hits[:limit]

    def read_zone(self, path: str, limit: int = 20) -> List[Dict]:
        """Return detailed rows from a zone workbook."""
        if not os.path.exists(path):
            return []
        out: List[Dict] = []
        with self._cross_process_lock():  # B4: exclude concurrent writers
            with self._lock:
                wb = load_workbook(path, read_only=True)
                ws = wb["mem"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    out.append({
                        "id": row[0], "title": row[1], "content": row[2],
                        "tags": row[3], "created": row[4],
                    })
        return out[-limit:]

    def list_zones(self) -> List[str]:
        zones = []
        for f in os.listdir(self.root):
            if f.endswith(".xlsx") and f != MASTER_NAME:
                zones.append(f[:-5])
        return zones

    # -- update / delete / forget (mutable memory) -------------------------

    def update(self, zone: str, row_id: int, brief: str = "", content: str = "",
               title: str = "", tags: str = "") -> bool:
        """Edit an existing memory row in a zone workbook + its master index entry.
        Returns True if something was changed. Wrapped in try/except so a failure
        is reported, not raised."""
        zone = zone or "knowledge"
        if zone not in ZONE_DEFAULTS:  # W1 (r8): enforce whitelist at public boundary
            return False
        try:
            with self._cross_process_lock():
                with self._lock:
                    zpath = self.zone_path(zone)
                    if not os.path.exists(zpath):
                        return False
                    # B2 (r7): before mutating the zone, verify the master index
                    # actually has this row for EVERY update (content-only too),
                    # not just when brief/title/tags are supplied. If the master
                    # row is missing the 1-master/1-zone invariant would break —
                    # refuse the update rather than leaving a zone-only row.
                    master_has_row = False
                    try:
                        mwb0 = load_workbook(self._master)
                        for mrow in mwb0["index"].iter_rows(min_row=2):
                            if mrow and mrow[0].value == row_id:
                                master_has_row = True
                                break
                    except Exception:
                        master_has_row = False
                    if not master_has_row:
                        return False
                    # Snapshot zone bytes for rollback if the master write fails
                    # (B5: keep zone and master consistent — no split-brain).
                    import shutil
                    z_snap = zpath + ".update.bak"
                    shutil.copy2(zpath, z_snap)
                    wb = load_workbook(zpath)
                    ws = wb["mem"]
                    changed = False
                    for row in ws.iter_rows(min_row=2):
                        if row and row[0].value == row_id:
                            if title:
                                row[1].value = self._safe_cell(title)
                            if content:
                                row[2].value = self._safe_cell(content)
                            if tags:
                                row[3].value = self._safe_cell(tags)
                            changed = True
                            break
                    # W3: if the zone row does not exist, do NOT touch master.
                    if not changed:
                        try:
                            os.remove(z_snap)
                        except OSError:
                            pass
                        return False
                    wb.save(zpath)
                    # also update the master index brief/title/tags if present
                    # (W2: tags-only updates must sync master tags too)
                    if brief or title or tags:
                        try:
                            mwb = load_workbook(self._master)
                            mws = mwb["index"]
                            master_hit = False
                            for mrow in mws.iter_rows(min_row=2):
                                if mrow and mrow[0].value == row_id:
                                    master_hit = True
                                    if brief:
                                        mrow[2].value = self._safe_cell(brief)
                                    if title:
                                        mrow[3].value = self._safe_cell(title)
                                    if tags:
                                        mrow[5].value = self._safe_cell(tags)
                                    break
                            if not master_hit:
                                # Zone row changed but the master index has no such
                                # row -> logical inconsistency (ChatGPT r6 BLOCKER):
                                # refuse the update and roll the zone back.
                                shutil.copy2(z_snap, zpath)
                                try:
                                    os.remove(z_snap)
                                except OSError:
                                    pass
                                return False
                            mwb.save(self._master)
                        except Exception:
                            # B5: master write failed -> roll zone back to snapshot
                            shutil.copy2(z_snap, zpath)
                            try:
                                os.remove(z_snap)
                            except OSError:
                                pass
                            return False
                    try:
                        os.remove(z_snap)
                    except OSError:
                        pass
                    return True
        except Exception as e:
            logging.getLogger(__name__).error("excel_line store.update failed: %s", e)
            return False

    def _remove_row(self, zone: str, row_id: int) -> bool:
        """Internal: delete one row from a zone workbook and the master index.
        Does NOT acquire the lock itself (caller must hold it).
        Atomic: snapshots the zone workbook before deleting; if the master index
        row is missing or its delete fails, the zone row is restored so
        master/zone stay consistent (ChatGPT round-6/7 BLOCKER: delete() must
        roll back when the master row is missing, not only when the master op throws)."""
        import shutil
        zpath = self.zone_path(zone)
        if not os.path.exists(zpath):
            return False
        # B1 (r7): before touching the zone, verify the master index actually has
        # this row. If the master row is missing, deleting only the zone would
        # break the 1-master/1-zone invariant — refuse and report no-op.
        master_has_row = False
        try:
            mwb0 = load_workbook(self._master)
            for mrow in mwb0["index"].iter_rows(min_row=2):
                if mrow and mrow[0].value == row_id:
                    master_has_row = True
                    break
        except Exception:
            master_has_row = False
        if not master_has_row:
            return False
        removed = False
        z_snap = zpath + ".delete.bak"
        shutil.copy2(zpath, z_snap)
        wb = load_workbook(zpath)
        ws = wb["mem"]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row and row[0].value == row_id:
                ws.delete_rows(i, 1)
                removed = True
                break
        if removed:
            wb.save(zpath)
        if removed:
            try:
                mwb = load_workbook(self._master)
                mws = mwb["index"]
                for i, mrow in enumerate(mws.iter_rows(min_row=2), start=2):
                    if mrow and mrow[0].value == row_id:
                        mws.delete_rows(i, 1)
                        break
                mwb.save(self._master)
            except Exception:
                shutil.copy2(z_snap, zpath)
                try:
                    os.remove(z_snap)
                except OSError:
                    pass
                return False
        try:
            os.remove(z_snap)
        except OSError:
            pass
        return removed
        try:
            os.remove(z_snap)
        except OSError:
            pass
        return removed

    def delete(self, zone: str, row_id: int) -> bool:
        """Delete a memory row from a zone workbook + master index. Returns True
        if a row was removed."""
        zone = zone or "knowledge"
        if zone not in ZONE_DEFAULTS:  # W1 (r8): enforce whitelist at public boundary
            return False
        try:
            with self._cross_process_lock():
                with self._lock:
                    return self._remove_row(zone, row_id)
        except Exception as e:
            logging.getLogger(__name__).error("excel_line store.delete failed: %s", e)
            return False

    def forget(self, query: str) -> int:
        """Delete every memory whose brief/title/tags match `query` (case-insensitive
        substring). Useful for 'forget everything about X'. Returns count removed."""
        q = (query or "").lower().strip()
        if not q:
            return 0
        try:
            with self._cross_process_lock():
                with self._lock:
                    removed = 0
                    # gather ids to delete per zone
                    wb = load_workbook(self._master, read_only=True)
                    ws = wb["index"]
                    targets = []  # (zone, id)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or row[0] is None:
                            continue
                        blob = " ".join(str(x or "") for x in (row[1], row[2], row[3], row[5])).lower()
                        if q in blob:
                            targets.append((row[1], row[0]))
                    for zone, rid in targets:
                        if self._remove_row(zone, rid):
                            removed += 1
                    return removed
        except Exception as e:
            logging.getLogger(__name__).error("excel_line store.forget failed: %s", e)
            return 0

    def count(self) -> int:
        try:
            with self._cross_process_lock():  # B4: exclude concurrent writers
                wb = load_workbook(self._master, read_only=True)
                return max(0, wb["index"].max_row - 1)
        except Exception:
            return 0

    # -- helpers -----------------------------------------------------------

    @staticmethod
    def _autosize(ws, cols):
        for i, _ in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 24
