"""migrate_v2b.py — dựng cây v2 trong <root>/brain/ mà KHÔNG đụng file v1 nào.

v1 giữ nguyên chỗ cũ (tự nhiên là bản archive sống). v2 tạo mới hoàn toàn:
  brain/brain.xlsx, brain/<zone>.xlsx, brain/<zone>/<tag>.xlsx …
Sau khi kiểm tra OK → đổi plugins.excel_line.root sang thư mục brain/ trong
config (hermes config set), plugin v2 đọc từ đó.

--run để thực thi.
"""
import os, re, sys, time
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))           # dữ liệu v1
BRAIN = os.path.join(HERE, "brain")                          # cây v2 mới
PLUGIN = r"C:\Users\Admin\AppData\Local\hermes\plugins\excel_line"
sys.path.insert(0, r"C:\Users\Admin\AppData\Local\hermes\plugins")
import types, importlib.util as _ilu, io
from openpyxl import load_workbook
_pkg = sys.modules.setdefault("excel_line", types.ModuleType("excel_line"))
_pkg.__path__ = [PLUGIN]
def _load_local(mod_name, file_name):
    spec = _ilu.spec_from_file_location(mod_name, os.path.join(PLUGIN, file_name))
    m = _ilu.module_from_spec(spec); sys.modules[mod_name] = m
    spec.loader.exec_module(m); return m
_bs = _load_local("excel_line.brain_store", "brain_store.py")
BrainStore, MASTER_V2, MAX_ROWS, FullError = _bs.BrainStore, _bs.MASTER_V2, _bs.MAX_ROWS, _bs.FullError

RUN = "--run" in sys.argv

def rows_of(path):
    with open(path, "rb") as f:
        data = f.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
    out = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)
           if r and r[0] is not None]
    wb.close()
    return out, hdr[0]

OLD_MASTER = os.path.join(HERE, "excel-line_index.xlsx")
if not os.path.exists(OLD_MASTER):
    print("không có v1 master."); sys.exit(0)
def v1_zone_path(z):
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in str(z).lower())
    return os.path.join(HERE, f"{safe}.xlsx")

master_rows, _ = rows_of(OLD_MASTER)
detail = {}
for z in {str(r["zone"]) for r in master_rows}:
    zp = v1_zone_path(z)
    if os.path.exists(zp):
        try:
            for r, _h in rows_of(zp):
                detail[int(r["id"])] = str(r.get("content") or "")
        except Exception:
            pass
by_zone = defaultdict(list)
for r in master_rows:
    by_zone[str(r["zone"] or "knowledge")].append(r)
print("v1:", len(master_rows), "dòng,", len(by_zone), "zone")
if not RUN:
    print("(dry-run — --run)"); sys.exit(0)

def sanitize(s):
    return re.sub(r"[^\w\-]", "_", str(s))[:24].strip("._") or "misc"
def first_tag(t):
    for part in (t or "").split(","):
        p = part.strip().lower()
        if p and p not in ("map-added", "auto-backup", "memory-md", "env",
                           "auto-extract", "classify-garbage", "qa", "live-test"):
            return sanitize(part.strip())
    return "misc"
def _chunks(lst, n):
    return [lst[i:i + n] for i in range(0, len(lst), n)]

os.makedirs(BRAIN, exist_ok=True)
for f in os.listdir(BRAIN):                    # fresh start nếu chạy lại
    fp = os.path.join(BRAIN, f)
    if fp.endswith(".xlsx"):
        os.remove(fp)
lock = os.path.join(BRAIN, ".excel_line.lock")
if os.path.exists(lock):
    os.remove(lock)
store = BrainStore(BRAIN)

warnings = []
for zone, items in sorted(by_zone.items(), key=lambda kv: -len(kv[1])):
    try:
        zres = store.child(MASTER_V2, sanitize(zone), f"📁 {zone}",
                           content=f"{len(items)} memories", tags="zone")
    except FullError:
        warnings.append("root đầy — bỏ zone " + zone); continue
    zfile = zres["file"]
    if len(items) <= MAX_ROWS:
        for r in items:
            store.add(zfile, title=(r.get("title") or r.get("brief") or "")[:50],
                      content=(detail.get(int(r["id"])) or r.get("brief") or "")[:250],
                      tags=r.get("tags") or "")
        continue
    groups = defaultdict(list)
    for r in items:
        groups[first_tag(r.get("tags"))].append(r)
    flat = []
    for tag, grows in sorted(groups.items(), key=lambda kv: -len(kv[1])):
        for i, part in enumerate(_chunks(grows, MAX_ROWS)):
            flat.append((tag if i == 0 else f"{tag}-{i+1}", part))
    # packing các group nhỏ chung node
    buf = []
    def flush():
        nonlocal_buf = buf[:]
        buf.clear()
        if not nonlocal_buf:
            return
        tags = sorted({first_tag(r.get("tags")) for r in nonlocal_buf})
        nm = tags[0] if len(tags) == 1 else "tạp-" + str(len(nonlocal_buf))
        try:
            sub = store.child(zfile, sanitize(nm), f"🏷 {nm} · {len(nonlocal_buf)}",
                              tags=",".join(tags))
            for r in nonlocal_buf:
                store.add(sub["file"], title=(r.get("title") or r.get("brief") or "")[:50],
                          content=(detail.get(int(r["id"])) or r.get("brief") or "")[:250],
                          tags=r.get("tags") or "")
        except FullError:
            store.add(zfile, title=f"nén {len(nonlocal_buf)} mục",
                      content=" | ".join((x.get("title") or x.get("brief") or "")[:40] for x in nonlocal_buf[:4])[:240],
                      tags="merged")
    master_get = lambda r, k: str(r.get(k) or "")
    for tag, rs in flat:
        if len(buf) + len(rs) <= MAX_ROWS:
            buf.extend(rs)
        else:
            flush(); buf.extend(rs)
            if len(buf) >= MAX_ROWS:
                flush()
    flush()

bad = []
def walk(b):
    rs = store.load_rows(b)
    if len(rs) > MAX_ROWS:
        bad.append((b, len(rs)))
    for r in rs:
        rb = str(r.get("branch") or "")
        if rb.lower().endswith(".xlsx"):
            walk(rb)
walk(MASTER_V2)
print("=== TREE v2 ===")
print(store.tree_text())
print("rows:", store.count(), "| vượt 10:", bad or "không", "| warnings:", warnings or "không")
