"""curate_v1.py — curation thủ công bằng LLM (main agent) trên 305 dòng còn lại.

Nguyên tắc đã chốt với user: chỉ giữ memory TÁI SỬ DỤNG ĐƯỢC.
- KEEP:  fact môi trường (#1-8,#16), hồ sơ user (#17-25,#117), skill custom (#122,#123)
- DISTILL: 6 business rule đào được từ chat-echo → viết thành memory chuẩn
- DELETE: toàn bộ chat echo còn lại (knowledge id khác)

Chạy dưới cross-process lock; backup .curate-*.bak; bulk block-delete (nhanh).
--run để thực thi.
"""
import os, shutil, sys, time
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, r"C:\Users\Admin\AppData\Local\hermes\plugins\excel_line")
from store import ExcelLineStore

store = ExcelLineStore(HERE)
RUN = "--run" in sys.argv

KEEP_KNOWLEDGE = set(range(1, 9)) | {16}
KEEP_USER = {17, 18, 19, 20, 21, 22, 23, 24, 25, 117}
KEEP_SKILL = {122, 123}

DISTILL = [
    ("knowledge", "Seagift: dòng không có tên khách trên hóa đơn = khách lẻ, KHÔNG lấy hóa đơn khi đối chiếu.",
     "Quy tắc đối chiếu bán hàng Seagift: những dòng không có tên là khách lẻ, không lấy hoá đơn. Người dùng đã khẳng định nhiều lần (26/08, 28/08).",
     "seagift,doi-soat,hoa-don"),
    ("knowledge", "Seagift CoA mapping: các mã khác → PHI01; phí gia công dùng mã 'CÁ GIA CÔNG HL'.",
     "Quy tắc mã hóa đơn/file Seagift: đổi các mã khác thành PHI01; dòng phí gia công mã đúng là 'CÁ GIA CÔNG HL'; file hóa đơn đối chiếu giữ đồng bộ theo mã này.",
     "seagift,coa,ma-hoa-don"),
    ("contact", "Đoàn Văn Mai: có thể vừa là nhà cung cấp vừa là khách hàng của Seagift.",
     "Khi gán mã khách/NCC tự động, Đoàn Văn Mai xuất hiện ở cả hai vai — không mặc định một chiều.",
     "seagift,doi-tac"),
    ("knowledge", "Hải sản: lịch từ 1kg trở lên đã tính là 'lịch lớn' — audit phải soi từng dòng.",
     "Ngưỡng nghiệp vụ: lượng lịch ≥1kg = lịch lớn; khi đối chiếu số lượng/đơn giá không được gộp lướt, từng dòng phải khớp SL×ĐG=GT.",
     "seagift,hai-san,audit"),
    ("pref", "Không lưu câu lệnh/trần thuật/hỏi đáp nhất thời của user vào memory — chỉ fact tái sử dụng được.",
     "Memory hygiene (user duyệt 03/09): input dạng 'làm X', 'sao bạn không Y', 'oke làm tiếp' KHÔNG phải memory. Sub-agent phải nén bằng LLM (brief ≤120 ký tự) như builtin memory, không copy nguyên văn.",
     "memory-hygiene,excel-line"),
    ("knowledge", "excel_line: classifier phải dùng ctx.llm (PluginLlm); LLM down → giữ log retry, cấm ghi nguyên văn.",
     "Bài học 03/09: core Hermes đã bỏ agent.run_agent.quick_completion → import fail nuốt lỗi trả JSON giả gây ô nhiễm 5.8k dòng. Fix: facade ctx.llm + tri-state ok/garbage/down; down = retry-log, không store. Đã push fork 477817b.",
     "excel-line,plugin,fix"),
]

def ranges_of(sorted_idx):
    out = []
    for i in sorted_idx:
        if out and i == out[-1][0] + out[-1][1]:
            out[-1][1] += 1
        else:
            out.append([i, 1])
    return [(s, c) for s, c in out]

# lock không reentrant: store.add()/delete() phải chạy NGOÀI khối lock của script
# 1. DISTILL trước (id mới > max cũ → không nằm trong diện xóa bên dưới)
new_ids = []
if RUN:
    for zone, brief, content, tags in DISTILL:
        mid = store.add(zone=zone, brief=brief, content=content,
                        title=brief[:40], tags=tags)
        new_ids.append(mid)
    print("distilled ids:", new_ids)

with store._cross_process_lock(timeout=30.0, stale_after=120.0):
    t0 = time.time()

    # 2. phân loại dòng master
    mwb = load_workbook(store._master)
    mws = mwb["index"]
    keep, dele = [], []
    for i, row in enumerate(mws.iter_rows(min_row=2), start=2):
        if not row or row[0].value is None:
            continue
        rid, zone = int(row[0].value), str(row[1].value)
        ok = (zone == "knowledge" and rid in KEEP_KNOWLEDGE) or \
             (zone == "user" and rid in KEEP_USER) or \
             (zone == "skill" and rid in KEEP_SKILL) or rid in new_ids
        (keep if ok else dele).append((i, zone, rid))
    print(f"keep: {len(keep)} | delete: {len(dele)}")
    if not RUN:
        print("(dry-run — thêm --run)")
        sys.exit(0)

    stamp = time.strftime("%Y%m%d-%H%M%S")
    shutil.copy2(store._master, store._master + f".curate-{stamp}.bak")
    for z in {d[1] for d in dele}:
        zp = store.zone_path(z)
        if os.path.exists(zp):
            shutil.copy2(zp, zp + f".curate-{stamp}.bak")

    for start, count in sorted(ranges_of([d[0] for d in dele]), reverse=True):
        mws.delete_rows(start, count)
    mwb.save(store._master)

    ids_by_zone = {}
    for _, z, rid in dele:
        ids_by_zone.setdefault(z, set()).add(rid)
    for z, ids in ids_by_zone.items():
        zp = store.zone_path(z)
        if not os.path.exists(zp):
            continue
        zwb = load_workbook(zp)
        zws = zwb["mem"]
        to_del = [i for i, row in enumerate(zws.iter_rows(min_row=2), start=2)
                  if row and row[0].value in ids]
        for start, count in sorted(ranges_of(to_del), reverse=True):
            zws.delete_rows(start, count)
        zwb.save(zp)

    wb = load_workbook(store._master, read_only=True)
    rows = [r for r in wb["index"].iter_rows(min_row=2, values_only=True) if r and r[0] is not None]
    wb.close()
    from collections import Counter
    print("after:", len(rows), dict(Counter(str(r[1]) for r in rows)),
          f"| {time.time()-t0:.1f}s")
