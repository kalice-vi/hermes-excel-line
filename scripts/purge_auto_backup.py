"""purge_auto_backup.py v2 — xóa dòng 'auto-backup' bằng bulk range delete.

(v1 chết vì gọi delete_rows 5.8k lần = O(n²) trong openpyxl)
Thuật toán: gom row index liên tiếp thành range, xóa từ range cuối ngược lên ->
số lần delete_rows = số block liên tiếp (thường 1), nhanh gần như tức thì.
"""
import os, shutil, sys, time
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, r"C:\Users\Admin\AppData\Local\hermes\plugins\excel_line")
from store import ExcelLineStore

store = ExcelLineStore(HERE)
DRY = "--run" not in sys.argv

def is_junk(tags):
    return tags == "auto-backup" or tags.startswith("auto-backup,llm-unavailable")

def ranges_of(sorted_idx):
    """[328,329,...,999] -> [(328,572)] các block liên tiếp (start, count)."""
    out = []
    for i in sorted_idx:
        if out and i == out[-1][0] + out[-1][1]:
            out[-1][1] += 1
        else:
            out.append([i, 1])
    return [(s, c) for s, c in out]

with store._cross_process_lock(timeout=30.0, stale_after=120.0):
    t0 = time.time()
    mwb = load_workbook(store._master)
    mws = mwb["index"]
    junk_rows = []      # row idx trong master
    junk_ids = {}       # zone -> set(ids)
    for i, row in enumerate(mws.iter_rows(min_row=2), start=2):
        if not row or row[0].value is None:
            continue
        if is_junk(str(row[5].value or "")):
            junk_rows.append(i)
            z = str(row[1].value)
            junk_ids.setdefault(z, set()).add(int(row[0].value))
    total = mws.max_row - 1
    print(f"junk: {len(junk_rows)} | tổng: {total}")

    if DRY:
        print("(dry-run — thêm --run để thực thi)")
        sys.exit(0)

    stamp = time.strftime("%Y%m%d-%H%M%S")
    for z in junk_ids:
        zp = store.zone_path(z)
        if os.path.exists(zp):
            shutil.copy2(zp, zp + f".purge-{stamp}.bak")
    shutil.copy2(store._master, store._master + f".purge-{stamp}.bak")

    # master: bulk delete từng block, từ cuối ngược lên
    for start, count in sorted(ranges_of(junk_rows), reverse=True):
        mws.delete_rows(start, count)
    mwb.save(store._master)

    # zones: chỉ những zone có junk, cũng bulk theo block id
    for z, ids in junk_ids.items():
        zp = store.zone_path(z)
        if not os.path.exists(zp):
            continue
        zwb = load_workbook(zp)
        zws = zwb["mem"]
        to_del = [i for i, row in enumerate(zws.iter_rows(min_row=2), start=2)
                  if row and row[0].value in ids]
        for start, count in sorted(ranges_of(to_del), reverse=True):
            zws.delete_rows(start, count)
        zwb.save(zp)

    # verify
    wb = load_workbook(store._master, read_only=True)
    rows = [r for r in wb["index"].iter_rows(min_row=2, values_only=True)
            if r and r[0] is not None]
    rem = sum(1 for r in rows if is_junk(str(r[5] or "")))
    wb.close()
    print(f"after: rows={len(rows)} | junk còn={rem} | {time.time()-t0:.1f}s")

    # orphan check: zone rows có id không còn trong master (còn trùng id cũ?)
    master_ids = {int(r[0]) for r in rows}
    for f in os.listdir(HERE):
        if f.endswith(".xlsx") and f != os.path.basename(store._master):
            z = f[:-5]
            if z == "excel-line_index": continue
            wb2 = load_workbook(os.path.join(HERE, f), read_only=True)
            orph = [r[0] for r in wb2["mem"].iter_rows(min_row=2, values_only=True)
                    if r and r[0] is not None and int(r[0]) not in master_ids]
            wb2.close()
            print(f"zone {z}: rows={len(master_ids and [1])*0 or ''}", end="")
            print(f" orphans={len(orph)} {orph[:10]}")
