"""migrate_v2.py — đưa excel_line v1 (master phẳng + zone files) lên cây v2.

AN TOÀN:
  - KHÔNG xóa file v1: copy sang <name>.v1archived-<stamp> rồi XÓA file gốc
    để cây v2 được dựng mới hoàn toàn, cùng tên (user.xlsx v2...).
  - Chỉ nhận diện v1 qua header chữ thường 'id'.
  - Mọi thao tác dưới cross-process lock của store.
  - --run để thực thi (mặc định dry-run).

Bố cục cây:
  brain.xlsx  : 1 dòng / zone (≤10 zone) → node <zone>.xlsx
  <zone>.xlsx : các memory của zone, nếu >9 thì mỗi nhóm tag là 1 node con
                <zone>/<tag>.xlsx (mỗi file ≤10, group >10 cắt -2, -3)
                — file zone chứa DÒNG LIÊN KẾT tới node con, không chứa trực
                tiếp khi đã tách nhánh (trừ khi vừa ≤10).
"""
import os, re, sys, time, shutil
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))           # thư mục dữ liệu
PLUGIN = r"C:\Users\Admin\AppData\Local\hermes\plugins\excel_line"
sys.path.insert(0, r"C:\Users\Admin\AppData\Local\hermes\plugins")
import types
import importlib.util as _ilu
from openpyxl import load_workbook
_pkg = sys.modules.setdefault("excel_line", types.ModuleType("excel_line"))
_pkg.__path__ = [PLUGIN]
def _load_local(mod_name, file_name):
    spec = _ilu.spec_from_file_location(mod_name, os.path.join(PLUGIN, file_name))
    m = _ilu.module_from_spec(spec); sys.modules[mod_name] = m
    spec.loader.exec_module(m); return m
_bs = _load_local("excel_line.brain_store", "brain_store.py")
BrainStore, MASTER_V2, MAX_ROWS, FullError = _bs.BrainStore, _bs.MASTER_V2, _bs.MAX_ROWS, _bs.FullError
_st = _load_local("excel_line.store", "store.py")
ExcelLineStore = _st.ExcelLineStore

RUN = "--run" in sys.argv
def v1_zone_path(z):
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in str(z).lower())
    return os.path.join(HERE, f"{safe}.xlsx")
OLD_MASTER = os.path.join(HERE, "excel-line_index.xlsx")
if not os.path.exists(OLD_MASTER):
    print("không tìm thấy v1 master — có thể đã migrate rồi."); sys.exit(0)

def rows_of(path):
    # đọc bytes rồi parse từ memory: không giữ file handle trên Windows
    import io
    with open(path, "rb") as f:
        data = f.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
    out = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)
           if r and r[0] is not None]
    wb.close()
    return out, hdr[0]

# 1. đọc v1
master_rows, _ = rows_of(OLD_MASTER)
detail = {}
for z in {str(r["zone"]) for r in master_rows}:
    zp = v1_zone_path(z)
    if os.path.exists(zp):
        try:
            for r, _h in rows_of(zp):
                detail[int(r["id"])] = str(r.get("content") or "")
        except Exception:
            pass
by_zone = defaultdict(list)
for r in master_rows:
    by_zone[str(r["zone"] or "knowledge")].append(r)
print("v1:", len(master_rows), "dòng,", len(by_zone), "zone")
for z, items in by_zone.items():
    print(f"   {z}: {len(items)}")
if not RUN:
    print("(dry-run — thêm --run)")
    sys.exit(0)

stamp = time.strftime("%Y%m%d-%H%M%S")
lock = os.path.join(HERE, ".excel_line.lock")
if os.path.exists(lock):
    os.remove(lock)

def sanitize(s):
    return re.sub(r"[^\w\-]", "_", str(s))[:24].strip("._") or "misc"

def first_tag(t):
    for part in (t or "").split(","):
        p = part.strip().lower()
        if p and p not in ("map-added", "auto-backup", "memory-md", "env",
                           "auto-extract", "classify-garbage", "qa", "live-test"):
            return sanitize(part.strip())
    return "misc"

# 2. archive + clear các file .xlsx v1
archived = []
for f in sorted(os.listdir(HERE)):
    if not f.endswith(".xlsx"):
        continue
    p = os.path.join(HERE, f)
    try:
        _r, h0 = rows_of(p)
    except Exception:
        continue
    if h0 == "id":                      # v1 layout (chữ thường)
        dst = p + f".v1archived-{stamp}"
        shutil.copy2(p, dst)
        os.remove(p)                    # xóa gốc để child() tạo file v2 mới tinh
        archived.append((f, os.path.basename(dst)))
print("archived v1:", [a[0] for a in archived])
# dọn phế phẩm v2 dở dang của lần chạy lỗi trước: brain.xlsx và mọi file
# header 'ID' đều là kết quả migrate hỏng → rename khỏi cây (.v2partial)
for f in list(os.listdir(HERE)):
    if not f.endswith(".xlsx"):
        continue
    p = os.path.join(HERE, f)
    try:
        rws, h0 = rows_of(p)
    except Exception:
        continue
    if h0 == "id":                       # còn sót v1 chưa kịp archive
        os.remove(p)
    elif h0 == "ID":                     # phế phẩm v2 hỏng → loại khỏi cây
        os.rename(p, p + f".v2partial-{stamp}")
        print("loại phế phẩm v2:", f)

# 3. dựng cây v2
store = BrainStore(HERE)
# BrainStore._ensure_master(v1) có thể tạo lại index rỗng — dọn
_stray = os.path.join(HERE, "excel-line_index.xlsx")
if os.path.exists(_stray):
    try:
        rws, h0 = rows_of(_stray)
        if h0 == "id" and not rws:
            os.remove(_stray)
    except Exception:
        pass

warnings = []

def link_rows(zfile, groups):
    """groups: list[(tagname, rows≤10)] → tạo node con, mỗi node 1 dòng link
    trong zfile (zfile chứa ≤10 dòng link)."""
    for nm, rs in groups:
        if len(store.load_rows(zfile)) >= MAX_ROWS:
            # zfile link đã đầy → gộp nốt groups còn lại vào 1 node 'tồn-overflow'
            rest = groups[groups.index((nm, rs)):]
            flat = [x for _g in rest for x in _g[1]]
            try:
                sub = store.child(zfile, "tồn", f"🏷 tồn · {len(flat)}", tags="overflow")
                for part in _chunks(flat, MAX_ROWS):
                    try:
                        f2 = store.child(sub["file"], f"t-{part[0]['id']}",
                                         f"{len(part)} mục", tags="overflow")
                        for r in part:
                            store.add(f2["file"], title=(r["title"] or r["brief"])[:50],
                                      content=(detail.get(int(r["id"])) or r["brief"])[:250],
                                      tags=r["tags"])
                    except FullError:
                        store.add(sub["file"], title=f"nén {len(part)} mục",
                                  content=" | ".join((x["title"] or x["brief"])[:40] for x in part[:4])[:240],
                                  tags="merged")
                return
            except FullError:
                warnings.append(f"{zfile} đầy, {len(flat)} dòng không migrate (xem archive)")
                return
        try:
            sub = store.child(zfile, nm, f"🏷 {nm} · {len(rs)}", tags=nm)
            for r in rs:
                store.add(sub["file"], title=(r["title"] or r["brief"])[:50],
                          content=(detail.get(int(r["id"])) or r["brief"])[:250],
                          tags=r["tags"])
        except FullError as e:
            warnings.append(str(e))

def _chunks(lst, n):
    return [lst[i:i + n] for i in range(0, len(lst), n)]

for zone, items in sorted(by_zone.items(), key=lambda kv: -len(kv[1])):
    zres = None
    try:
        zres = store.child(MASTER_V2, sanitize(zone), f"📁 {zone}",
                           content=f"{len(items)} memories (migrate v1)", tags="zone")
    except FullError:
        warnings.append("root đầy — zone " + zone + " bỏ lại")
        continue
    zfile = zres["file"]
    if len(items) <= MAX_ROWS:
        for r in items:
            store.add(zfile, title=(r["title"] or r["brief"])[:50],
                      content=(detail.get(int(r["id"])) or r["brief"])[:250],
                      tags=r["tags"])
        continue
    groups = defaultdict(list)
    for r in items:
        groups[first_tag(r["tags"])].append(r)
    # cắt group >10
    flat_groups = []
    for tag, grows in sorted(groups.items(), key=lambda kv: -len(kv[1])):
        if len(grows) <= MAX_ROWS:
            flat_groups.append((tag, grows))
        else:
            parts = _chunks(grows, MAX_ROWS)
            for i, pt in enumerate(parts):
                flat_groups.append((tag if i == 0 else f"{tag}-{i + 1}", pt))
    # gộp các group nhỏ vào nhau ≤10 để đỡ phình số node
    packed, buf = [], []
    for tag, rs in flat_groups:
        if len(buf) + len(rs) <= MAX_ROWS:
            buf += rs
        else:
            packed.append(("gộp-" + (first_tag(buf[0]["tags"]) if buf else "n"), buf))
            buf = list(rs)
    if buf:
        packed.append(("gộp-" + (first_tag(buf[0]["tags"]) if buf else "n"), buf))
    link_rows(zfile, packed)

# 4. kiểm bất biến + tổng kết
bad = []
def walk(b):
    rs = store.load_rows(b)
    if len(rs) > MAX_ROWS:
        bad.append((b, len(rs)))
    for r in rs:
        rb = str(r.get("branch") or "")
        if rb.lower().endswith(".xlsx"):
            walk(rb)
try:
    walk(MASTER_V2)
except Exception as e:
    warnings.append("walk: " + str(e))
print("\n=== TREE v2 ===")
print(store.tree_text())
print("rows v2:", store.count(), "| file vượt 10:", bad or "không", "| warnings:", warnings or "không")
