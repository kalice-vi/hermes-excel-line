"""brain_server.py — server mind-map đồng bộ 2 chiều THẲNG vào excel_line.

- GET  /api/data     : render cây mới nhất từ Excel (không qua cache file)
- GET  /api/search   : ?q=... tìm theo brief/tag/id
- POST /api/ops      : [{op...}, ...] áp trực tiếp vào ExcelLineStore (có lock)
- tĩnh               : phục vụ editor.html + js/css

Ngôn ngữ chung: xem brain_map.py. Ops được thiết kế idempotent-ish và
CHỈ đi qua store — không bao giờ sửa .xlsx bằng đường vòng.

Memory do user tạo trên map mang tag 'map-added' để agent nhận diện
(không tự xử lý/làm lại phần user đã gõ).
"""
from __future__ import annotations
import json, os, sys, threading, urllib.parse
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)  # brain_map
STATIC = r"C:\Users\Admin\AppData\Local\Temp\mindmap"
PLUGIN = r"C:\Users\Admin\AppData\Local\hermes\plugins\excel_line"
sys.path.insert(0, PLUGIN)

from brain_map import build_tree, topic_of  # noqa: E402
from mermaid_map import build_mmd, diff_ops  # noqa: E402
from store import ExcelLineStore, ZONE_DEFAULTS  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

ROOT = os.environ.get("EXCEL_LINE_ROOT", HERE)
store = ExcelLineStore(ROOT)
_lock = threading.Lock()

# ------------------------------------------------------------------ helpers

def zone_of(rid: int):
    """Tìm (zone, row) trong master index theo id."""
    wb = load_workbook(store._master, read_only=True, data_only=True)
    try:
        for r in wb["index"].iter_rows(min_row=2, values_only=True):
            if r and r[0] == rid:
                return {"zone": str(r[1]), "brief": str(r[2] or ""),
                        "title": str(r[3] or ""), "tags": str(r[5] or "")}
    finally:
        wb.close()
    return None

def content_of(rid: int, zone: str) -> str:
    p = store.zone_path(zone)
    if not os.path.exists(p):
        return ""
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        for r in wb["mem"].iter_rows(min_row=2, values_only=True):
            if r and r[0] == rid:
                return str(r[2] or "")
    finally:
        wb.close()
    return ""

def set_refs(rid: int, zone: str, refs):
    """Ghi/đổi mảng REFS: trong content (zone workbook) qua store.update."""
    content = content_of(rid, zone)
    import re
    line = "REFS: " + "; ".join(refs) if refs else None
    if re.search(r"REFS:.+", content):
        new = re.sub(r"REFS:.+", line or "REFS:", content)
    else:
        new = (content + ("\n" if content else "") + line) if line else content
    return store.update(zone, rid, content=new)

# ------------------------------------------------------------------ ops

def op_add(o):
    zone = o.get("zone") or "knowledge"
    if zone not in ZONE_DEFAULTS:
        return f"zone '{zone}' không hợp lệ"
    topic = (o.get("topic") or "New").strip()[:120]
    tags = "map-added" + ("," + topic_of(o.get("tag")) if o.get("tag") else "")
    mid = store.add(zone, brief=topic, content=topic, title=topic, tags=tags)
    return {"id": mid} if mid > 0 else "store.add trả -1"

def op_rename(o):
    rid = int(o["id"]); z = zone_of(rid)
    if not z: return "không tìm thấy #" + str(rid)
    ok = store.update(z["zone"], rid, brief=o["topic"], title=o["topic"][:40])
    return {"ok": ok} if ok else "update thất bại"

def op_retag(o):
    rid = int(o["id"]); z = zone_of(rid)
    if not z: return "không tìm thấy #" + str(rid)
    tags = "map-added," + (o.get("tag") or "").strip() if "map-added" in z["tags"] else (o.get("tag") or "").strip()
    ok = store.update(z["zone"], rid, tags=tags)
    return {"ok": ok} if ok else "update thất bại"

def op_move(o):
    """Đổi zone (kéo memory sang nhánh zone khác) hoặc đổi tag (kéo sang topic)."""
    rid = int(o["id"]); z = zone_of(rid)
    if not z: return "không tìm thấy #" + str(rid)
    new_zone = o.get("zone"); new_tag = o.get("tag")
    if new_zone and new_zone in ZONE_DEFAULTS:
        # add dòng zone mới + xóa dòng zone cũ (giữ nội dung)
        content = content_of(rid, z["zone"])
        mid2 = store.add(new_zone, brief=z["brief"], content=content,
                         title=z["title"], tags=z["tags"])
        if mid2 > 0:
            store.delete(z["zone"], rid)
            return {"moved": True, "new_id": mid2}
        return "add zone mới thất bại"
    if new_tag is not None:
        old = [t.strip() for t in z["tags"].split(",") if t.strip()]
        keep = [t for t in old if t.startswith("map-added")]
        newtags = ",".join(keep + ([new_tag] if new_tag else []))
        ok = store.update(z["zone"], rid, tags=newtags)
        return {"ok": ok} if ok else "update tags thất bại"
    return "thiếu zone/tag"

def op_delete(o):
    rid = int(o["id"]); z = zone_of(rid)
    if not z: return "không tìm thấy #" + str(rid)
    ok = store.delete(z["zone"], rid)
    return {"ok": ok} if ok else "delete thất bại"

def op_addref(o):
    rid = int(o["id"]); z = zone_of(rid)
    if not z: return "không tìm thấy #" + str(rid)
    path = (o.get("path") or "").strip()
    if not path: return "thiếu path"
    refs = []
    import re
    m = re.search(r"REFS:(.+)", content_of(rid, z["zone"]))
    if m:
        refs = [x.strip() for x in m.group(1).split(";") if x.strip()]
    if path in refs:
        return {"ok": True, "exists": True}
    refs.append(path)
    ok = set_refs(rid, z["zone"], refs)
    return {"ok": ok} if ok else "ghi refs thất bại"

def op_delref(o):
    rid = int(o["id"]); z = zone_of(rid)
    if not z: return "không tìm thấy #" + str(rid)
    import re
    m = re.search(r"REFS:(.+)", content_of(rid, z["zone"]))
    refs = [x.strip() for x in m.group(1).split(";") if x.strip()] if m else []
    path = (o.get("path") or "").strip()
    refs = [r for r in refs if r != path]
    ok = set_refs(rid, z["zone"], refs)
    return {"ok": ok} if ok else "xóa refs thất bại"

OPS = {"add": op_add, "rename": op_rename, "retag": op_retag, "move": op_move,
       "delete": op_delete, "addref": op_addref, "delref": op_delref}

def apply_ops(ops):
    results = []
    for o in ops:
        fn = OPS.get(o.get("op"))
        if not fn:
            results.append({"op": o, "error": "op lạ: " + str(o.get("op"))})
            continue
        try:
            r = fn(o)
        except Exception as e:
            r = f"exception: {e}"
        if isinstance(r, str):
            r = {"error": r}
        results.append({"op": o.get("op"), "id": o.get("id"), **r})
    return results

# ------------------------------------------------------------------ http

class H(BaseHTTPRequestHandler):
    def log_message(self, *a): pass

    def _send(self, code, body, ctype="application/json"):
        data = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", ctype + "; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        u = urllib.parse.urlparse(self.path)
        if u.path == "/api/data":
            try:
                # build_tree đọc read-only, lock do store quản lý khi ghi —
                # KHÔNG giữ cross-process lock ở đây (non-reentrant, xem mmd_sync)
                tree = build_tree(ROOT)
                idx = os.path.join(ROOT, "excel-line_index.xlsx")
                tree["_rev"] = os.stat(idx).st_mtime if os.path.exists(idx) else 0
            except Exception as e:
                return self._send(500, json.dumps({"error": str(e)}, ensure_ascii=False))
            return self._send(200, json.dumps(tree, ensure_ascii=False))
        if u.path == "/api/search":
            q = urllib.parse.parse_qs(u.query).get("q", [""])[0]
            hits = store.search_index(q, limit=50) if q else []
            return self._send(200, json.dumps({"hits": hits}, ensure_ascii=False))
        if u.path == "/api/mmd":
            # .mmd sinh trực tiếp từ Excel + revision để client phát hiện đổi
            import hashlib
            try:
                with store._cross_process_lock(timeout=5.0):
                    mmd = build_mmd(ROOT)
            except Exception as e:
                return self._send(500, json.dumps({"error": str(e)}, ensure_ascii=False))
            rev = hashlib.md5(mmd.encode("utf-8")).hexdigest()[:12]
            return self._send(200, json.dumps({"rev": rev, "mmd": mmd}, ensure_ascii=False))
        if u.path == "/api/mmd_sync":
            pass  # POST only
        # static (editor + assets) — ưu tiên TEMP/mindmap, fallback thư mục plugin
        p = u.path.lstrip("/") or "editor.html"
        for base in (STATIC, HERE):
            fp = os.path.normpath(os.path.join(base, p))
            if fp.startswith(base) and os.path.isfile(fp):
                ctype = {".html": "text/html", ".js": "text/javascript",
                         ".css": "text/css", ".map": "application/json",
                         ".mmd": "text/plain"}.get(
                    os.path.splitext(fp)[1], "application/octet-stream")
                with open(fp, "rb") as f:
                    return self._send(200, f.read(), ctype)
        return self._send(404, b"not found", "text/plain")

    def do_POST(self):
        path = urllib.parse.urlparse(self.path).path
        if path == "/api/ops":
            try:
                n = int(self.headers.get("Content-Length", 0))
                payload = json.loads(self.rfile.read(n).decode("utf-8"))
                ops = payload if isinstance(payload, list) else [payload]
                with _lock:
                    results = apply_ops(ops)
                return self._send(200, json.dumps({"results": results}, ensure_ascii=False))
            except Exception as e:
                return self._send(400, json.dumps({"error": str(e)}, ensure_ascii=False))
        if path == "/api/mmd_sync":
            # client gửi cả văn bản .mmd đã sửa -> server tự diff với Excel -> ops
            try:
                n = int(self.headers.get("Content-Length", 0))
                payload = json.loads(self.rfile.read(n).decode("utf-8"))
                new_mmd = payload.get("mmd", "")
                # chỉ giữ _lock của server; KHÔNG giữ cross-process lock ở đây
                # vì store.add/update/delete sẽ tự lấy lock (nó không reentrant)
                with _lock:
                    old_mmd = build_mmd(ROOT)
                    ops = diff_ops(old_mmd, new_mmd)
                    results = apply_ops(ops) if ops else []
                    final_mmd = build_mmd(ROOT)
                return self._send(200, json.dumps(
                    {"applied": len(ops), "results": results, "mmd": final_mmd},
                    ensure_ascii=False))
            except Exception as e:
                return self._send(400, json.dumps({"error": str(e)}, ensure_ascii=False))
        return self._send(404, '{"error":"?"}')

if __name__ == "__main__":
    port = int(os.environ.get("BRAIN_PORT", "8766"))
    print(f"brain server → http://127.0.0.1:{port}/editor.html (root={ROOT})")
    ThreadingHTTPServer(("127.0.0.1", port), H).serve_forever()
