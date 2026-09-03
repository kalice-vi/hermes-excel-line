"""project_map.py — Chiếu (projection) excel_line -> cây jsMind.

Chiều duy nhất ĐƯỢC PHÉP đọc tự động: Excel -> Map. Map luôn là "cửa sổ",
Excel là sự thật (source of truth). Agent/user muốn đổi map -> gửi ops,
ops được áp vào Excel qua store có lock, rồi map tự reload từ Excel.

Cây:
  🧠 BRAIN (user, ngày sync)
  ├── 📁 Zone: knowledge (312)
  │     └── 🏷 topic-tag nhóm (nếu có)
  │           └── 📄 brief  [row #id]
  │                 └── 🔗 file/tài nguyên (nếu link có trong tags/content)
  ├── 📁 Zone: skill ...
  └── 📁 Zone: user ...
"""
import json, os, re, sys, time
from openpyxl import load_workbook

ROOT = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    "~\\AppData\\Local\\hermes\\excel_line")
MASTER = os.path.join(ROOT, "excel-line_index.xlsx")

LINK_RE = re.compile(r"(?:[A-Za-z]:[\\/][^\s,;\"']+|\\\\[^\s,;]+|[/~][A-Za-z0-9_.\-/\\ ]{6,})")

def rows():
    wb = load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb["index"]
    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0] is not None:
            out.append(dict(zip(header, r)))
    wb.close()
    return out

def topic_of(tags):
    t = (tags or "").strip()
    if not t:
        return "♾ Other"
    first = t.split(",")[0].strip()
    return first or "♾ Other"

def build():
    data = rows()
    zones = {}
    for row in data:
        z = str(row.get("zone") or "misc").strip() or "misc"
        zones.setdefault(z, []).append(row)

    children = []
    for zi, (zone, items) in enumerate(sorted(zones.items(), key=lambda kv: -len(kv[1]))):
        topics = {}
        for row in items:
            topics.setdefault(topic_of(row.get("tags")), []).append(row)
        topic_nodes = []
        for topic, trows in sorted(topics.items(), key=lambda kv: -len(kv[1])):
            mem_nodes = []
            for row in trows:
                blob = " ".join(str(row.get(k) or "") for k in ("path", "tags", "brief"))
                links = []
                selfpath = str(row.get("path") or "").replace("\\\\", "\\")
                for m in set(LINK_RE.findall(blob)):
                    p = m.strip().rstrip('.,)')
                    if p.replace("\\\\", "\\").lower() == selfpath.lower():
                        continue  # bản thân file zone không phải "link"
                    if len(p) >= 6 and p not in links:
                        links.append(p)
                children_of_mem = [
                    {"id": f"l{row['id']}h{abs(hash(p)) % 99999}",
                     "topic": "🔗 " + (os.path.basename(p.replace("\\\\", "\\")) or p)}
                    for p in links[:10]]
                brief = str(row.get("brief") or "").replace("\n", " ")
                mem_nodes.append({
                    "id": f"m{row['id']}",
                    "topic": f"📄 {brief[:90]} [#{row['id']}]",
                    "children": children_of_mem})
            topic_nodes.append({"id": f"t{zi}{abs(hash(topic)) % 99999}",
                                "topic": f"🏷 {topic} ({len(mem_nodes)})",
                                "children": mem_nodes})
        children.append({"id": f"z{zi}", "topic": f"📁 {zone} ({len(items)})",
                         "children": topic_nodes})

    tree = {"id": "root",
            "topic": f"🧠 BRAIN · {len(data)} memories · {time.strftime('%Y-%m-%d %H:%M')}",
            "children": children}
    payload = {"meta": {"name": "brain", "author": "excel_line", "version": "0.9.1"},
               "format": "node_tree", "data": tree}
    with open(os.path.join(ROOT, "brain.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    print("zones:", {z: len(v) for z, v in zones.items()}, "total:", len(data))

if __name__ == "__main__":
    build()
