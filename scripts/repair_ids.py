"""repair_ids.py — vá lỗi id TRÙNG trong master index + zone workbooks.

Nguyên nhân: store._new_id() đọc max(id) từ file tại thời điểm ghi; hai writer
xen kẽ (worker + provider) cùng đọc được max cũ -> cùng cấp 1 id. Hệ quả:
80 cặp (master row + zone row) dùng chung id.

Chiến thuật (an toàn, chạy dưới cross-process lock của store):
- Giữ nguyên bản đầu tiên của mỗi id; các bản sau được cấp id mới (max+1...).
- Cặp master[i] <-> zone[i] khớp theo THỨ TỰ XUẤT HIỆN (cùng lúc add nên
  cùng thứ tự trong file).
- Backup .bak từng workbook trước khi sửa; chỉ ghi khi có thay đổi.
- Idempotent: chạy lần 2 không đổi gì.
"""
import os, shutil, sys, time

HERE = os.path.dirname(os.path.abspath(__file__))
PLUGIN = r"C:\Users\Admin\AppData\Local\hermes\plugins\excel_line"
sys.path.insert(0, PLUGIN)
from openpyxl import load_workbook          # noqa: E402
from store import ExcelLineStore             # noqa: E402

store = ExcelLineStore(HERE)

with store._cross_process_lock(timeout=30.0, stale_after=60.0):
    # 1. quét master, tìm dup ids
    mwb = load_workbook(store._master)
    mws = mwb["index"]
    seen = {}          # id -> number of occurrences seen
    max_id = 0
    dup_rows = []      # (excel_row_idx, old_id) rows needing renumber
    for i, row in enumerate(mws.iter_rows(min_row=2), start=2):
        v = row[0].value
        if isinstance(v, int):
            max_id = max(max_id, v)
            seen[v] = seen.get(v, 0) + 1
            if seen[v] > 1:
                dup_rows.append((i, v))
    print("master rows:", mws.max_row - 1, "| dup occurrences:", len(dup_rows))
    if not dup_rows:
        print("nothing to repair")
        sys.exit(0)

    # 2. với mỗi zone: đếm số lần xuất hiện id trong sheet mem
    zone_cursors = {}  # zone -> {id: occurrence index}
    renum = []         # (zone, old_id, new_id, master_row_idx)
    need_zone = {}
    for row_idx, old_id in dup_rows:
        zone = str(mws.cell(row=row_idx, column=2).value)
        need_zone.setdefault(zone, []).append((row_idx, old_id))

    # kiểm tra zone workbook có đủ số bản để đối chiếu theo thứ tự
    for zone, items in need_zone.items():
        zpath = store.zone_path(zone)
        if not os.path.exists(zpath):
            print(f"⚠ zone workbook missing: {zpath} — skip zone")
            continue
        zwb = load_workbook(zpath)
        zws = zwb["mem"]
        zcount = {}
        for row in zws.iter_rows(min_row=2, values_only=True):
            if row and isinstance(row[0], int):
                zcount[row[0]] = zcount.get(row[0], 0) + 1
        zc = zone_cursors.setdefault(zone, {})
        for row_idx, old_id in items:
            zc[old_id] = zc.get(old_id, 0) + 1
            zocc = zc[old_id]
            if zocc > zcount.get(old_id, 0):
                print(f"⚠ zone {zone}: chỉ có {zcount.get(old_id,0)} bản id={old_id}"
                      f" nhưng master có ≥{zocc} — bản {zocc} chỉ sửa master")
            max_id += 1
            renum.append((zone, old_id, max_id, row_idx, zocc))

    # 3. backup + ghi zone workbooks
    touched = set()
    backups = []
    for zone, old_id, new_id, _, zocc in renum:
        touched.add(zone)
    stamp = time.strftime("%Y%m%d-%H%M%S")
    for zone in touched:
        zpath = store.zone_path(zone)
        if not os.path.exists(zpath):
            continue
        b = zpath + f".idfix-{stamp}.bak"
        shutil.copy2(zpath, b)
        backups.append(b)
    b = store._master + f".idfix-{stamp}.bak"
    shutil.copy2(store._master, b)
    backups.append(b)

    zwb_cache = {}
    for zone, old_id, new_id, _, zocc in renum:
        zpath = store.zone_path(zone)
        if not os.path.exists(zpath):
            continue
        zwb = zwb_cache.get(zone) or load_workbook(zpath)
        zwb_cache[zone] = zwb
        n = 0
        for row in zwb["mem"].iter_rows(min_row=2):
            if row and row[0].value == old_id:
                n += 1
                if n == zocc:
                    row[0].value = new_id
                    row[5].value = store and time.strftime("%Y-%m-%d %H:%M:%S")
                    break

    for zone, wb in zwb_cache.items():
        wb.save(store.zone_path(zone))

    # 4. ghi master
    for zone, old_id, new_id, row_idx, _ in renum:
        mws.cell(row=row_idx, column=1).value = new_id
        mws.cell(row=row_idx, column=8).value = time.strftime("%Y-%m-%d %H:%M:%S")
    mwb.save(store._master)

    # 5. verify
    wb = load_workbook(store._master, read_only=True)
    ids = [r[0] for r in wb["index"].iter_rows(min_row=2, values_only=True)
           if r and isinstance(r[0], int)]
    wb.close()
    from collections import Counter
    dups = {k: v for k, v in Counter(ids).items() if v > 1}
    print("after: rows =", len(ids), "unique =", len(set(ids)), "dups =", len(dups))
    print("backups:", [os.path.basename(x) for x in backups])
