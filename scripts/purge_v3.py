"""purge v3 — xóa các dòng echo RÁC HỆ THỐNG còn sót từ các bản lỗi trước:
 - 'You are a memory classifier…'   (prompt classifier bị echo lại)
 - '[CONTEXT COMPACTION …]'         (tóm tắt nén context, không phải memory)
 - '[System: …'                     (thông báo hệ thống giữa phiên)
 - '[/learn] …'                     (prompt skill bị echo)
 - 'QA probe …'                     (dòng test)
Chạy dưới cross-process lock, có backup. --run để thực thi.
"""
import os, re, shutil, sys, time
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, r"C:\Users\Admin\AppData\Local\hermes\plugins\excel_line")
from store import ExcelLineStore

store = ExcelLineStore(HERE)
DRY = "--run" not in sys.argv

JUNK = re.compile(
    r"^(You are a memory classifier"
    r"|\[CONTEXT COMPACTION"
    r"|\[System:"
    r"|\[/learn\]"
    r"|QA probe"
    r"|T\u00f4i vừa mua một chiếc xe m\u00e1y Honda Vision 2026"  # dữ liệu test cũ
    r")", re.I)

def ranges_of(sorted_idx):
    out = []
    for i in sorted_idx:
        if out and i == out[-1][0] + out[-1][1]:
            out[-1][1] += 1
        else:
            out.append([i, 1])
    return [(s, c) for s, c in out]

with store._cross_process_lock(timeout=30.0, stale_after=120.0):
    mwb = load_workbook(store._master)
    mws = mwb["index"]
    junk_rows, junk_ids = [], {}
    for i, row in enumerate(mws.iter_rows(min_row=2), start=2):
        if not row or row[0].value is None:
            continue
        if JUNK.search(str(row[2].value or "")):
            junk_rows.append(i)
            junk_ids.setdefault(str(row[1].value), set()).add(int(row[0].value))
    print("junk system-echo:", len(junk_rows))
    if DRY or not junk_rows:
        print("(dry-run hoặc không có gì để xóa)"); sys.exit(0)
    stamp = time.strftime("%Y%m%d-%H%M%S")
    for z in junk_ids:
        zp = store.zone_path(z)
        if os.path.exists(zp):
            shutil.copy2(zp, zp + f".purge3-{stamp}.bak")
    shutil.copy2(store._master, store._master + f".purge3-{stamp}.bak")
    for start, count in sorted(ranges_of(junk_rows), reverse=True):
        mws.delete_rows(start, count)
    mwb.save(store._master)
    for z, ids in junk_ids.items():
        zp = store.zone_path(z)
        if not os.path.exists(zp):
            continue
        zwb = load_workbook(zp)
        zws = zwb["mem"]
        to_del = [i for i, row in enumerate(zws.iter_rows(min_row=2), start=2)
                  if row and row[0].value in ids]
        for start, count in sorted(ranges_of(to_del), reverse=True):
            zws.delete_rows(start, count)
        zwb.save(zp)
    wb = load_workbook(store._master, read_only=True)
    rows = [r for r in wb["index"].iter_rows(min_row=2, values_only=True) if r and r[0] is not None]
    wb.close()
    from collections import Counter
    print("after:", len(rows), dict(Counter(str(r[1]) for r in rows)))
