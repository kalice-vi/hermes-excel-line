"""mermaid_map.py — ngôn ngữ chung text giữa excel_line và Mermaid mindmap.

Quy ước ánh xạ (mỗi dòng .mmd là một node, indent = cấp bậc).
KHÔNG dùng ( ) [ ] trong text node vì Mermaid đó là cú pháp shape:

    mindmap
      root((🧠 BRAIN))
        📁 knowledge 222             <- zone
          🏷 env 42                   <- tag nhóm
            📄119 brief...           <- memory (id sát sau 📄)
              🔗119i0 filename.xlsx  <- ref file

Chiều Excel -> .mmd : build_mmd(root) — chiếu thuần, format chuẩn hóa.
Chiều .mmd -> Excel : diff_ops(old,new) — suy ra ops, server ghi qua store.
Memory user thêm từ pane mang tag 'map-added' (agent không đụng tới).
"""
from __future__ import annotations
import os, re, difflib
from typing import Dict, List, Tuple

from openpyxl import load_workbook

IND = "  "   # 2 space / cấp

# ---------------------------------------------------------------- Excel -> mmd

def read_index(root: str) -> List[Dict]:
    p = os.path.join(root, "excel-line_index.xlsx")
    if not os.path.exists(p):
        return []
    wb = load_workbook(p, read_only=True, data_only=True)
    header = [str(c.value or "").strip().lower() for c in next(wb["index"].iter_rows(max_row=1))]
    rows = [dict(zip(header, r)) for r in wb["index"].iter_rows(min_row=2, values_only=True)
            if r and r[0] is not None]
    wb.close()
    return rows

def read_refs(root: str, zones) -> Dict[int, List[str]]:
    out: Dict[int, List[str]] = {}
    for z in zones:
        p = os.path.join(root, z + ".xlsx")
        if not os.path.exists(p):
            continue
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
            for r in wb["mem"].iter_rows(min_row=2, values_only=True):
                if r and r[0] is not None:
                    m = re.search(r"REFS:(.+)", str(r[2] or ""))
                    if m:
                        out[int(r[0])] = [x.strip() for x in m.group(1).split(";") if x.strip()]
            wb.close()
        except Exception:
            pass
    return out

def topic_of(tags) -> str:
    t = (str(tags or "")).strip()
    for part in t.split(","):
        part = part.strip()
        if part and part != "map-added" and not part.startswith(("memory-md", "env")):
            return part
    return "♾ khác"

def _clean(s: str) -> str:
    """Thoát ký tự phá cú pháp mindmap + gộp về 1 dòng."""
    s = re.sub(r"\s+", " ", str(s))
    return re.sub(r"[\(\)\[\]{}<>#&`;]", " ", s).strip()

def build_mmd(root: str) -> str:
    rows = read_index(root)
    refs = read_refs(root, sorted({str(r.get("zone") or "misc") for r in rows}))
    zones: Dict[str, List[Dict]] = {}
    for row in rows:
        zones.setdefault(str(row.get("zone") or "misc"), []).append(row)
    lines = ["mindmap", f"{IND}root((🧠 BRAIN · {len(rows)} memories))"]
    for zone, items in sorted(zones.items(), key=lambda kv: -len(kv[1])):
        lines.append(f"{IND*2}📁 {_clean(zone)} · {len(items)}")
        topics: Dict[str, List[Dict]] = {}
        for row in items:
            topics.setdefault(topic_of(row.get("tags")), []).append(row)
        for topic, trows in sorted(topics.items(), key=lambda kv: -len(kv[1])):
            lines.append(f"{IND*3}🏷 {_clean(topic)} · {len(trows)}")
            for row in trows:
                rid = int(row["id"])
                brief = _clean(str(row.get("brief") or ""))[:80]
                lines.append(f"{IND*4}📄{rid} {brief}")
                for i, pp in enumerate(refs.get(rid, [])[:12]):
                    lines.append(f"{IND*5}🔗{rid}i{i} {_clean(os.path.basename(pp))}")
    return "\n".join(lines) + "\n"

# ---------------------------------------------------------------- mmd -> ops

MEM_RE = re.compile(r"📄(\d+)\s+(.*)")
NEW_RE = re.compile(r"📄\s+(\S.*)")     # user gõ: 📄 không kèm id
ZONE_RE = re.compile(r"📁\s*(\S+)")
TAG_RE = re.compile(r"🏷\s*(.+?)\s*·\s*\d+$")

def _parse(text: str) -> Dict[str, Dict]:
    """id -> {zone, tag, brief}. Nhận diện theo EMOJI đầu dòng + số khoảng trắng."""
    out: Dict[str, Dict] = {}
    zone = tag = None
    for line in text.splitlines():
        stripped = line.lstrip()
        if not stripped:
            continue
        indent = len(line) - len(stripped)
        if stripped.startswith("📁") and indent <= 4:
            m = ZONE_RE.search(line); zone = m.group(1) if m else zone; tag = None
        elif stripped.startswith("🏷") and indent <= 6:
            m = TAG_RE.search(line)
            if m: tag = m.group(1).strip()
        elif stripped.startswith("📄") and indent >= 6:
            m = MEM_RE.match(stripped)
            if m:
                rid = int(m.group(1))
                entry = {"zone": zone or "knowledge", "tag": tag or "",
                         "brief": m.group(2).strip()}
                if rid in out:
                    # id TRÙNG trong master index (lỗi dữ liệu excel_line):
                    # node thứ 2+ mang key 'dup:N' — chỉ đọc, không ops được
                    n = 2
                    while f"dup:{rid}:{n}" in out:
                        n += 1
                    out[f"dup:{rid}:{n}"] = entry
                else:
                    out[rid] = entry
            else:
                m2 = NEW_RE.match(stripped)
                if m2:
                    out["new:" + m2.group(1).strip()] = {
                        "zone": zone or "knowledge", "tag": tag or "",
                        "brief": m2.group(1).strip()}
    return out

def diff_ops(old: str, new: str) -> List[Dict]:
    a, b = _parse(old), _parse(new)
    ops: List[Dict] = []
    for rid in a:
        if isinstance(rid, str):   # 'dup:…' — node ma, bỏ qua mọi ops
            continue
        if rid not in b:
            ops.append({"op": "delete", "id": rid})
        else:
            if a[rid]["brief"] != b[rid]["brief"]:
                ops.append({"op": "rename", "id": rid, "topic": b[rid]["brief"]})
            if a[rid]["zone"] != b[rid]["zone"]:
                ops.append({"op": "move", "id": rid, "zone": b[rid]["zone"]})
            elif a[rid]["tag"] != b[rid]["tag"]:
                ops.append({"op": "move", "id": rid, "tag": b[rid]["tag"]})
    for rid in b:
        if rid not in a:
            ops.append({"op": "add", "zone": b[rid]["zone"], "topic": b[rid]["brief"],
                        "tag": b[rid]["tag"], "map_new_id": rid})
    return ops

if __name__ == "__main__":
    r = os.environ.get("EXCEL_LINE_ROOT", ".")
    mmd = build_mmd(r)
    open(os.path.join(r, "brain.mmd"), "w", encoding="utf-8").write(mmd)
    print("lines:", mmd.count("\n"))
    # self-test diff rỗng: parse(build) phải cho đúng số memory
    print("memories parsed:", len(_parse(mmd)))
