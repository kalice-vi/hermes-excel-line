"""brain_map.py — NGÔN NGỮ CHUNG giữa Excel-line và Mind map.

Quy ước ID node (cả server lẫn editor đều hiểu):
    root              : nút gốc
    z:<zone>          : nhánh zone (ví dụ z:knowledge)
    t:<zone>:<tag>    : nhánh topic = tag đầu của memory
    m:<row_id>        : một memory (row_id = id trong master index)
    r:<row_id>:<n>    : một 🔗 tham chiếu file của memory đó
    new:...           : node tạm do editor tạo (sẽ thành m:<id> sau khi ghi)

Chiều dữ liệu:
    Excel (store, có lock) --render--> Map        :luôn luôn
    Map --ops (add/rename/retag/move/delete/addref)--> store   :mọi sửa đổi
    Memory do user thêm trên map mang tag 'map-added'
    => agent chỉ xử lý memory MỚI do chính nó add, không đụng map-added.

REFS lưu trong cột content của sheet mem, dạng dòng cuối:
    REFS: D:\\a.xlsx; C:\\b.md   (mảng 'ref;' phân tách bằng '; ')
"""
from __future__ import annotations
import os, re, sys, time
from typing import Dict, List

from openpyxl import load_workbook

LINK_RE = re.compile(r"(?:[A-Za-z]:[\\/][^\s,;\"']+|\\\\[^\s,;]+|/~?[A-Za-z0-9_.\-/\\ ]{6,})")

def read_index(root: str) -> List[Dict]:
    p = os.path.join(root, "excel-line_index.xlsx")
    if not os.path.exists(p):
        return []
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb["index"]
    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0] is not None:
            rows.append(dict(zip(header, r)))
    wb.close()
    return rows

def read_refs(root: str, zones: List[str]) -> Dict[int, List[str]]:
    """row_id -> danh sách REFS (từ content) + path nhắc trong brief."""
    out: Dict[int, List[str]] = {}
    for z in zones:
        p = os.path.join(root, z + ".xlsx")
        if not os.path.exists(p):
            continue
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
            for r in wb["mem"].iter_rows(min_row=2, values_only=True):
                if not r or r[0] is None:
                    continue
                rid, content = r[0], str(r[2] or "")
                refs: List[str] = []
                m = re.search(r"REFS:(.+)", content)
                if m:
                    refs += [x.strip() for x in m.group(1).split(";") if x.strip()]
                out[int(rid)] = refs
            wb.close()
        except Exception:
            pass
    return out

def topic_of(tags) -> str:
    t = (str(tags or "")).strip()
    for part in t.split(","):
        part = part.strip()
        if part and part != "map-added" and not part.startswith(("memory-md", "env")):
            return part
    return "♾ khác"

def build_tree(root: str) -> Dict:
    """Chiếu Excel -> cây jsMind (đọc không, có lock của store qua read_only)."""
    rows = read_index(root)
    refs = read_refs(root, sorted({str(r.get("zone") or "misc") for r in rows}))
    zones: Dict[str, List[Dict]] = {}
    for row in rows:
        zones.setdefault(str(row.get("zone") or "misc"), []).append(row)

    znodes = []
    for zone, items in sorted(zones.items(), key=lambda kv: -len(kv[1])):
        topics: Dict[str, List[Dict]] = {}
        for row in items:
            topics.setdefault(topic_of(row.get("tags")), []).append(row)
        tnodes = []
        for topic, trows in sorted(topics.items(), key=lambda kv: -len(kv[1])):
            mnodes = []
            for row in trows:
                rid = int(row["id"])
                links = list(refs.get(rid, []))  # chỉ nhận REFS tường minh (addref trên map)
                rnodes = [{"id": f"r:{rid}:{i}",
                           "topic": "🔗 " + (os.path.basename(pp.replace("\\\\", "\\")) or pp)}
                          for i, pp in enumerate(links[:12])]
                brief = str(row.get("brief") or "").replace("\n", " ")
                mnodes.append({"id": f"m:{rid}",
                               "topic": f"📄 {brief[:90]} [#{rid}]",
                               "children": rnodes})
            tnodes.append({"id": f"t:{zone}:{topic}", "topic": f"🏷 {topic} ({len(mnodes)})",
                           "children": mnodes})
        znodes.append({"id": f"z:{zone}", "topic": f"📁 {zone} ({len(items)})",
                       "children": tnodes})

    total = sum(len(v) for v in zones.values())
    tree = {"id": "root",
            "topic": f"🧠 BRAIN · {total} memories · {time.strftime('%H:%M:%S')}",
            "children": znodes}
    return {"meta": {"name": "brain", "author": "excel_line", "version": "0.9.1"},
            "format": "node_tree", "data": tree}

if __name__ == "__main__":
    import json
    root = sys.argv[1] if len(sys.argv) > 1 else "."
    t = build_tree(root)
    print(json.dumps(t["data"]["topic"]))
    n = 0
    for z in t["data"]["children"]:
        for tp in z["children"]:
            n += len(tp["children"])
    print("leaf memories:", n)
